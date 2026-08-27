from io import BytesIO

from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from scheduling.exports.common import POSITION_CODES, assumptions, show_export_rows
from scheduling.models import Employee
from scheduling.services.metrics import metrics_for_employee

NAVY = "173B63"
BLUE = "2F78B8"
PALE_BLUE = "EAF2F8"
WHITE = "FFFFFF"


def _style_sheet(sheet, widths: dict[int, int]) -> None:
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.row_dimensions[1].height = 32
    for index, width in widths.items():
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        if row[0].row % 2 == 0:
            for cell in row:
                cell.fill = PatternFill("solid", fgColor=PALE_BLUE)


def build_schedule_workbook(schedule_run) -> Workbook:
    workbook = Workbook()
    schedule = workbook.active
    schedule.title = "Schedule"
    schedule.append(
        [
            "Date",
            "Day",
            "Show",
            "Expected Guests",
            "Lead Server",
            "Server 2",
            "Server 3",
            "On-Call Server",
            "Bartender",
            "On-Call Bartender",
            "Busser",
            "50/50",
            "Warnings",
        ]
    )
    for show, assignments, warnings in show_export_rows(schedule_run):
        people = [
            assignments.get(code).employee.display_name if assignments.get(code) else "SHORTAGE"
            for code in POSITION_CODES
        ]
        schedule.append(
            [
                show.date,
                show.date.strftime("%A"),
                show.title,
                show.planning_guest_count,
                *people,
                "\n".join(warning.get_warning_type_display() for warning in warnings),
            ]
        )
    for cell in schedule["A"][1:]:
        cell.number_format = "yyyy-mm-dd"
    _style_sheet(
        schedule,
        {1: 13, 2: 12, 3: 32, 4: 16, **{index: 20 for index in range(5, 13)}, 13: 35},
    )

    detailed = workbook.create_sheet("Detailed Assignments")
    detailed.append(
        [
            "Date",
            "Show",
            "Employee",
            "Role",
            "Assignment Type",
            "Start",
            "End",
            "Paid Hours",
            "On-Call Hours",
            "Selection Reason",
        ]
    )
    assignments = schedule_run.assignments.select_related(
        "show", "employee", "role", "shift_template"
    )
    for assignment in assignments:
        local_start = timezone.localtime(assignment.start_datetime)
        local_end = timezone.localtime(assignment.end_datetime)
        detailed.append(
            [
                assignment.show.date,
                assignment.show.title,
                assignment.employee.display_name,
                assignment.role.name,
                assignment.get_assignment_type_display(),
                local_start.strftime("%H:%M"),
                local_end.strftime("%H:%M"),
                float(assignment.scheduled_paid_hours),
                float(assignment.on_call_hours),
                assignment.selection_reason,
            ]
        )
    for cell in detailed["A"][1:]:
        cell.number_format = "yyyy-mm-dd"
    _style_sheet(
        detailed,
        {1: 13, 2: 30, 3: 22, 4: 15, 5: 18, 6: 10, 7: 10, 8: 12, 9: 14, 10: 70},
    )

    totals = workbook.create_sheet("Employee Totals")
    totals.append(
        [
            "Employee",
            "Confirmed Hours",
            "Confirmed Shifts",
            "On-Call Assignments",
            "Server Shifts",
            "Bartender Shifts",
            "Busser Shifts",
            "50/50 Shifts",
            "Weekend Shifts",
        ]
    )
    for employee in Employee.objects.filter(active=True):
        metric = metrics_for_employee(employee, schedule_run)
        totals.append(
            [
                employee.display_name,
                float(metric.confirmed_paid_hours),
                metric.confirmed_shift_count,
                metric.on_call_assignment_count,
                metric.server_shift_count,
                metric.bartender_shift_count,
                metric.busser_shift_count,
                metric.fifty_fifty_count,
                metric.weekend_assignment_count,
            ]
        )
    _style_sheet(totals, {1: 24, **{index: 20 for index in range(2, 10)}})

    warnings = workbook.create_sheet("Warnings")
    warnings.append(["Date", "Show", "Severity", "Type", "Message", "Resolved", "Resolution"])
    for warning in schedule_run.warnings.select_related("show"):
        warnings.append(
            [
                warning.show.date if warning.show else "Schedule period",
                warning.show.title if warning.show else "All shows",
                warning.get_severity_display(),
                warning.get_warning_type_display(),
                warning.message,
                "Yes" if warning.resolved else "No",
                warning.resolution_note,
            ]
        )
    _style_sheet(warnings, {1: 16, 2: 30, 3: 20, 4: 30, 5: 70, 6: 12, 7: 45})

    assumption_sheet = workbook.create_sheet("Assumptions")
    assumption_sheet.append(["Topic", "Detail"])
    for topic, detail in assumptions(schedule_run):
        assumption_sheet.append([topic, detail])
    _style_sheet(assumption_sheet, {1: 24, 2: 100})
    return workbook


def schedule_workbook_bytes(schedule_run) -> bytes:
    output = BytesIO()
    build_schedule_workbook(schedule_run).save(output)
    return output.getvalue()
