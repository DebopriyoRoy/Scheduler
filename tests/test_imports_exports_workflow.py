import re
from datetime import date, time
from io import BytesIO

import pytest
from django.contrib.auth import get_user_model
from django.core.exceptions import ValidationError
from django.core.management import call_command
from django.test import override_settings
from django.utils import timezone
from openpyxl import load_workbook

from scheduling.exports.csv_export import detailed_schedule_csv
from scheduling.exports.excel import schedule_workbook_bytes
from scheduling.exports.pdf_export import schedule_pdf_bytes
from scheduling.importers.availability import (
    AvailabilityCSVError,
    import_availability_rows,
    parse_availability_csv,
)
from scheduling.importers.calendar import SpiritCalendarImporter, events_from_html
from scheduling.models import (
    AvailabilityType,
    Employee,
    EmployeeAvailability,
    ScheduleRunStatus,
    SchedulingWarning,
    Show,
    WarningSeverity,
    WarningType,
)
from scheduling.services.engine import SchedulingEngine
from scheduling.services.workflow import approve_schedule, override_assignment


@pytest.fixture
def staff_and_config(db):
    call_command("seed_spirit_staff", verbosity=0)
    call_command("seed_scheduling_config", verbosity=0)
    return list(Employee.objects.filter(active=True))


@pytest.fixture
def management_user(db):
    return get_user_model().objects.create_user(username="manager", password="test-password")


def generated_run(staff, *, requires_50_50=True):
    # 80 guests is the standard band: 3 confirmed servers, 1 on-call, 1 bartender +
    # 1 on-call, 1 busser, 1 50/50 - eight positions on an ordinary night.
    show = Show.objects.create(
        title="Forever Country",
        date=date(2026, 9, 12),
        expected_guests=80,
        requires_50_50=requires_50_50,
    )
    for employee in staff:
        EmployeeAvailability.objects.create(
            employee=employee,
            date=show.date,
            availability_type=AvailabilityType.AVAILABLE_ALL_DAY,
        )
    return SchedulingEngine().generate(show.date, show.date), show


@pytest.mark.django_db
def test_availability_csv_validates_then_imports_atomically(staff_and_config):
    csv_text = (
        "employee,date,available,start_time,end_time,notes\n"
        "Olena,2026-09-12,yes,15:00,23:00,Can work\n"
        "Kate,2026-09-12,no,,,Away\n"
    )
    rows = parse_availability_csv(csv_text)
    assert EmployeeAvailability.objects.count() == 0
    assert import_availability_rows(rows) == 2
    olena = EmployeeAvailability.objects.get(employee__display_name="Olena")
    kate = EmployeeAvailability.objects.get(employee__display_name="Kate")
    assert olena.availability_type == AvailabilityType.AVAILABLE_WINDOW
    assert kate.availability_type == AvailabilityType.UNAVAILABLE


@pytest.mark.django_db
def test_malformed_availability_csv_imports_nothing(staff_and_config):
    csv_text = (
        "employee,date,available,start_time,end_time,notes\n"
        "Olena,2026-09-12,yes,,,Fine\n"
        "Not A Person,2026-09-12,yes,,,Bad\n"
    )
    with pytest.raises(AvailabilityCSVError, match="unknown employee"):
        parse_availability_csv(csv_text)
    assert EmployeeAvailability.objects.count() == 0


def test_calendar_parser_reads_json_ld_event():
    html = """
    <html><script type="application/ld+json">
    {"@type":"Event","name":"Forever Country","startDate":"2026-09-12T18:30:00-02:30",
    "endDate":"2026-09-12T22:30:00-02:30","url":"https://example.test/forever",
    "location":{"@type":"Place","name":"Theatre Gower"}}
    </script></html>
    """
    events = events_from_html(html, "https://example.test/calendar")
    assert len(events) == 1
    assert events[0].title == "Forever Country"
    assert events[0].date == date(2026, 9, 12)
    assert str(events[0].start_time) == "18:30:00"


def test_calendar_parser_has_human_visible_date_fallback():
    html = "<html><h1>Shift Happens</h1><p>September 19, 2026 6:30 pm – 10:30 pm</p></html>"
    events = events_from_html(html, "https://example.test/shift-happens")
    assert [(event.title, event.date) for event in events] == [("Shift Happens", date(2026, 9, 19))]


@pytest.mark.django_db
def test_calendar_import_is_idempotent_and_does_not_delete_manual_show():
    manual = Show.objects.create(title="Manual", date=date(2026, 9, 20))
    html = """
    <html><script type="application/ld+json">
    {"@type":"Event","name":"Imported","startDate":"2026-09-12T18:30:00-02:30",
    "endDate":"2026-09-12T22:30:00-02:30","url":"https://example.test/imported"}
    </script></html>
    """

    class Response:
        text = html

        def raise_for_status(self):
            return None

    class Session:
        headers = {}

        def get(self, *args, **kwargs):
            return Response()

    importer = SpiritCalendarImporter(session=Session())
    first = importer.import_range(date(2026, 9, 7), date(2026, 10, 3))
    second = importer.import_range(date(2026, 9, 7), date(2026, 10, 3))
    assert first.created == 1
    assert second.created == 0
    assert Show.objects.filter(pk=manual.pk).exists()


@pytest.mark.django_db
def test_excel_csv_and_pdf_exports_are_complete(staff_and_config):
    run, show = generated_run(staff_and_config)
    workbook_bytes = schedule_workbook_bytes(run)
    workbook = load_workbook(BytesIO(workbook_bytes), data_only=False)
    assert workbook.sheetnames == [
        "Schedule",
        "Detailed Assignments",
        "Employee Totals",
        "Warnings",
        "Assumptions",
    ]
    assert workbook["Schedule"]["A2"].value.date() == show.date
    assert workbook["Schedule"]["C2"].value == show.title
    # header + 9 assignments: the six-plus-two crew plus the Server Manager.
    assert workbook["Detailed Assignments"].max_row == 10
    detailed_rows = list(workbook["Detailed Assignments"].iter_rows(values_only=True))
    # Lead Server comes in 45 minutes before this show's doors (18:30) and leaves
    # 15 minutes after wrap (22:30).
    lead_row = next(row for row in detailed_rows if row[3] == "Server" and row[5] == "15:00")
    assert lead_row[6] == "21:00"
    # header + 18 staff, the roster having gained the Server Manager.
    assert workbook["Employee Totals"].max_row == 19
    assert "Forever Country" in detailed_schedule_csv(run)
    pdf = schedule_pdf_bytes(run)
    assert pdf.startswith(b"%PDF")
    assert len(pdf) > 3000


@pytest.mark.django_db
def test_the_exported_rota_carries_every_position_and_its_hours(staff_and_config):
    """A printed rota has to say when people come in, not just who is on.

    Both exports listed names alone and left the Server Manager column out entirely,
    so a printed sheet showed one fewer person than the schedule on screen and none of
    the call times the whole timetable exists to communicate.
    """
    from scheduling.exports.common import cell_text

    run, _show = generated_run(staff_and_config)
    sheet = load_workbook(BytesIO(schedule_workbook_bytes(run)), data_only=False)["Schedule"]

    headers = [cell.value for cell in sheet[1]]
    assert headers[4:13] == [
        "Server Manager",
        "Lead Server",
        "Server 2",
        "Server 3",
        "On-Call Server",
        "Bartender",
        "On-Call Bartender",
        "Busser",
        "50/50",
    ]

    # Every filled cell reads "Name" then the hours that person actually works.
    filled = 0
    for cell in sheet[2][4:13]:
        value = cell.value
        if value == "SHORTAGE":
            continue
        name, _, hours = value.partition("\n")
        assert name.strip(), f"no name in {value!r}"
        assert re.fullmatch(r"\d{2}:\d{2}-\d{2}:\d{2}", hours), f"no hours in {value!r}"
        filled += 1
    assert filled >= 7

    # The hours come from the assignment, so a trimmed shift prints its real hours.
    manager = run.assignments.get(shift_template__code="server-manager")
    assert cell_text(manager) == (
        f"{manager.employee.display_name}\n"
        f"{timezone.localtime(manager.start_datetime):%H:%M}-"
        f"{timezone.localtime(manager.end_datetime):%H:%M}"
    )
    assert cell_text(None) == "SHORTAGE"


@pytest.mark.django_db
def test_valid_override_is_audited_and_invalid_override_is_blocked(staff_and_config):
    run, _ = generated_run(staff_and_config, requires_50_50=False)
    assignment = run.assignments.get(shift_template__code="lead-server")
    replacement = next(
        employee
        for employee in Employee.objects.filter(employee_roles__role__name="Server")
        if not run.assignments.filter(show=assignment.show, employee=employee).exists()
    )
    result = override_assignment(assignment, replacement, "Management requested this change")
    assert result.manually_overridden is True
    assert result.override_reason == "Management requested this change"
    unavailable = Employee.objects.get(display_name="Olena")
    EmployeeAvailability.objects.filter(employee=unavailable, date=assignment.show.date).update(
        availability_type=AvailabilityType.UNAVAILABLE,
        available=False,
    )
    with pytest.raises(ValidationError, match="ineligible"):
        override_assignment(result, unavailable, "Try an invalid unavailable person")


@pytest.mark.django_db
def test_approval_blocks_errors_and_records_success(staff_and_config, management_user):
    run, _ = generated_run(staff_and_config)
    warning = SchedulingWarning.objects.create(
        schedule_run=run,
        warning_type=WarningType.SERVER_SHORTAGE,
        severity=WarningSeverity.ERROR,
        message="Test hard error",
    )
    with pytest.raises(ValidationError, match="Resolve"):
        approve_schedule(run, management_user)
    warning.resolved = True
    warning.resolution_note = "Management verified alternate coverage."
    warning.save()
    approve_schedule(run, management_user)
    run.refresh_from_db()
    assert run.status == ScheduleRunStatus.APPROVED
    assert run.approved_by == management_user
    assert run.approved_at is not None


@pytest.mark.django_db
def test_management_pages_and_export_endpoints(client, management_user, staff_and_config):
    run, _ = generated_run(staff_and_config)
    client.force_login(management_user)
    for url in (
        "/shows/",
        "/availability/",
        "/configuration/rotations/",
        "/schedules/",
        "/schedules/generate/",
        f"/schedules/{run.pk}/",
    ):
        assert client.get(url).status_code == 200
    excel_response = client.get(f"/schedules/{run.pk}/export.xlsx")
    assert excel_response.status_code == 200
    assert "spreadsheetml" in excel_response["Content-Type"]
    assert client.get(f"/schedules/{run.pk}/export.csv").status_code == 200
    assert client.get(f"/schedules/{run.pk}/export.pdf").content.startswith(b"%PDF")


@pytest.mark.django_db
def test_schedule_generation_page_requires_explicit_shortage_choice(
    client,
    management_user,
    staff_and_config,
):
    Show.objects.create(title="Missing Availability", date=date(2026, 9, 12))
    client.force_login(management_user)
    response = client.post(
        "/schedules/generate/",
        {"start_date": "2026-09-07", "end_date": "2026-10-03"},
    )
    assert response.status_code == 200
    assert b"Choose Generate with shortages" in response.content


@pytest.mark.django_db
def test_complete_browser_workflow_generates_overrides_and_approves(
    client,
    management_user,
    staff_and_config,
):
    show = Show.objects.create(
        title="Browser Workflow Show",
        date=date(2026, 9, 12),
        expected_guests=100,
        requires_50_50=True,
    )
    for employee in staff_and_config:
        EmployeeAvailability.objects.create(
            employee=employee,
            date=show.date,
            availability_type=AvailabilityType.AVAILABLE_ALL_DAY,
        )
    client.force_login(management_user)
    response = client.post(
        "/schedules/generate/",
        {"start_date": "2026-09-07", "end_date": "2026-10-03"},
    )
    assert response.status_code == 302
    schedule_run = show.schedule_assignments.first().schedule_run
    assert schedule_run.status == ScheduleRunStatus.GENERATED
    detail = client.get(response["Location"])
    assert detail.status_code == 200
    assert b"Browser Workflow Show" in detail.content
    assignment = schedule_run.assignments.get(shift_template__code="lead-server")
    replacement = next(
        employee
        for employee in Employee.objects.filter(employee_roles__role__name="Server")
        if not schedule_run.assignments.filter(show=show, employee=employee).exists()
    )
    override_response = client.post(
        f"/schedules/assignments/{assignment.pk}/override/",
        {"employee": replacement.pk, "override_reason": "Browser workflow validation"},
    )
    assert override_response.status_code == 302
    assignment.refresh_from_db()
    assert assignment.manually_overridden is True
    approval_response = client.post(f"/schedules/{schedule_run.pk}/approve/")
    assert approval_response.status_code == 302
    schedule_run.refresh_from_db()
    assert schedule_run.status == ScheduleRunStatus.APPROVED


@pytest.mark.django_db
@override_settings(DEBUG=True)
def test_demo_seed_is_idempotent_and_clearly_isolated():
    call_command("seed_schedule_demo", verbosity=0)
    call_command("seed_schedule_demo", verbosity=0)
    assert Show.objects.count() == 6
    assert not Show.objects.exclude(source=Show.Source.DEMO).exists()
    assert EmployeeAvailability.objects.count() == 108
    assert not EmployeeAvailability.objects.exclude(source="DEMO").exists()


@pytest.mark.django_db
def test_import_adopts_an_existing_show_when_the_external_id_changes():
    """A changed calendar id must update the night, not add a second copy of it.

    The show list had grown three rows per date because the same real show arrived
    under three different external_id schemes over time - a demo seed, an older
    title-slug importer, and the current one - and keying only on external_id made
    each of them a fresh row.
    """
    existing = Show.objects.create(
        title="Forever Country",
        date=date(2026, 9, 12),
        start_time=time(18, 30),
        end_time=time(22, 0),
        external_id="spirit-occ-forever-country-old-scheme",
        source=Show.Source.CALENDAR_IMPORT,
        active=True,
    )

    html = """
    <html><script type="application/ld+json">
    {"@type":"Event","name":"Forever Country... in the Key of Spirit!!",
    "startDate":"2026-09-12T18:30:00-02:30","endDate":"2026-09-12T22:30:00-02:30",
    "url":"https://example.test/occ-2026-09-12"}
    </script></html>
    """

    class Response:
        text = html

        def raise_for_status(self):
            return None

    class Session:
        headers = {}

        def get(self, *args, **kwargs):
            return Response()

    summary = SpiritCalendarImporter(session=Session()).import_range(
        date(2026, 9, 7), date(2026, 10, 3)
    )

    assert summary.created == 0, "a duplicate row was created for a night already held"
    assert summary.updated == 1
    assert Show.objects.filter(date=date(2026, 9, 12)).count() == 1

    existing.refresh_from_db()
    assert existing.title == "Forever Country... in the Key of Spirit!!"
    assert existing.end_time == time(22, 30)
    assert existing.external_id != "spirit-occ-forever-country-old-scheme"


@pytest.mark.django_db
def test_import_still_separates_two_shows_on_the_same_date():
    """Adoption keys on the start time, so a matinee and an evening show stay distinct."""
    matinee = Show.objects.create(
        title="Ugly Stick Workshop",
        date=date(2026, 9, 12),
        start_time=time(13, 0),
        end_time=time(15, 0),
        external_id="workshop-2026-09-12",
        source=Show.Source.CALENDAR_IMPORT,
        active=True,
    )

    html = """
    <html><script type="application/ld+json">
    {"@type":"Event","name":"Forever Country","startDate":"2026-09-12T18:30:00-02:30",
    "endDate":"2026-09-12T22:30:00-02:30","url":"https://example.test/evening"}
    </script></html>
    """

    class Response:
        text = html

        def raise_for_status(self):
            return None

    class Session:
        headers = {}

        def get(self, *args, **kwargs):
            return Response()

    SpiritCalendarImporter(session=Session()).import_range(
        date(2026, 9, 7), date(2026, 10, 3)
    )

    assert Show.objects.filter(date=date(2026, 9, 12)).count() == 2
    matinee.refresh_from_db()
    assert matinee.title == "Ugly Stick Workshop"


@pytest.mark.django_db
def test_replacing_with_someone_already_on_the_show_moves_them(staff_and_config):
    """Putting a person on a second position means moving them, not cloning them.

    This used to be refused outright - "already assigned another role for this show" -
    which is the engine arguing with a decision the manager has already made.
    """
    from scheduling.models import SchedulingWarning
    from scheduling.services.workflow import override_assignment

    run, show = generated_run(staff_and_config)
    target = run.assignments.get(shift_template__code="bartender")
    # Somebody already on this show who can actually hold the bar - the on-call
    # bartender assists there and is barred from the confirmed position by design.
    mover = next(
        row
        for row in run.assignments.exclude(pk=target.pk).filter(show=show)
        if row.employee.employee_roles.filter(
            role__name="Bartender", active=True, on_call_only=False
        ).exists()
    )
    moved_person, old_slot = mover.employee, mover.shift_template.name

    override_assignment(target, moved_person, "covering the bar tonight")

    target.refresh_from_db()
    assert target.employee == moved_person
    # They hold the new position and only the new position.
    assert run.assignments.filter(show=show, employee=moved_person).count() == 1
    assert not run.assignments.filter(pk=mover.pk).exists()
    assert old_slot in target.selection_reason
    assert "now unfilled" in target.selection_reason

    # The slot they left is reported, not silently dropped.
    gap = SchedulingWarning.objects.filter(
        schedule_run=run, show=show, severity=WarningSeverity.ERROR
    ).order_by("-id").first()
    assert gap is not None
    assert old_slot in gap.message
    assert moved_person.display_name in gap.message


@pytest.mark.django_db
def test_a_move_still_has_to_pass_the_real_checks(staff_and_config):
    """Moving somebody is not a way around time off, qualification or availability."""
    from scheduling.models import EmployeeTimeOff, TimeOffStatus
    from scheduling.services.workflow import override_assignment

    run, show = generated_run(staff_and_config)
    target = run.assignments.get(shift_template__code="bartender")
    mover = next(
        row
        for row in run.assignments.exclude(pk=target.pk).filter(show=show)
        if row.employee.employee_roles.filter(
            role__name="Bartender", active=True, on_call_only=False
        ).exists()
    )
    EmployeeTimeOff.objects.create(
        employee=mover.employee,
        start_date=show.date,
        end_date=show.date,
        status=TimeOffStatus.APPROVED,
        reason="Out of province",
    )

    with pytest.raises(ValidationError, match="Approved time off"):
        override_assignment(target, mover.employee, "covering the bar tonight")

    # Nothing moved, and their original shift is untouched.
    assert run.assignments.filter(pk=mover.pk).exists()


@pytest.mark.django_db
def test_two_people_on_one_show_can_swap_positions(staff_and_config):
    """Trading two shifts is one decision, so it is one action.

    Without it the only route was to empty one position, fill the other, then remember
    to come back - with a hole in the roster in between.
    """
    from scheduling.models import SchedulingWarning
    from scheduling.services.workflow import override_assignment

    run, show = generated_run(staff_and_config)
    first = run.assignments.get(shift_template__code="lead-server")
    second = run.assignments.get(shift_template__code="server-2")
    first_person, second_person = first.employee, second.employee
    first_window = (first.start_datetime, first.end_datetime)
    second_window = (second.start_datetime, second.end_datetime)

    override_assignment(first, second_person, "trading the two server shifts", swap=True)

    first.refresh_from_db()
    # The row for the second position is rebuilt rather than updated - both uniqueness
    # rules are database constraints and a straight swap breaks them mid-write - so it
    # is re-read by position, not by id.
    second = run.assignments.get(shift_template__code="server-2")
    assert first.employee == second_person
    assert second.employee == first_person
    # Each takes the hours of the position, not of the person.
    assert (first.start_datetime, first.end_datetime) == first_window
    assert (second.start_datetime, second.end_datetime) == second_window
    # A swap leaves no hole, so nothing is reported as vacated.
    assert not SchedulingWarning.objects.filter(
        schedule_run=run, show=show, message__contains="was moved"
    ).exists()


@pytest.mark.django_db
def test_someone_who_only_assists_in_a_role_never_holds_it(staff_and_config):
    """Neil backs the bartender up; he does not open or close the bar himself."""
    from scheduling.models import EmployeeRole
    from scheduling.services.workflow import override_assignment

    run, _show = generated_run(staff_and_config)
    assistant = EmployeeRole.objects.filter(
        role__name="Bartender", on_call_only=True, active=True
    ).first()
    assert assistant is not None, "the seed should mark the bar assistant on-call only"

    # Never rostered to the confirmed position by the engine...
    confirmed = run.assignments.filter(shift_template__code="bartender").first()
    if confirmed:
        assert confirmed.employee != assistant.employee

    # ...and not placeable there by hand either.
    with pytest.raises(ValidationError, match="on-call support only"):
        override_assignment(confirmed, assistant.employee, "cover the bar tonight")


@pytest.mark.django_db
def test_a_bar_restriction_does_not_follow_someone_onto_the_floor(staff_and_config):
    """Neil assists at the bar but is an ordinary server.

    The restriction is held per role, so qualifying him for a second role leaves the
    first one's limit exactly where it was. If it were held per person, putting him on
    the floor would either drag the bar limit with it or quietly lift it.
    """
    from scheduling.models import EmployeeRole, Role, ScheduleRun, ShiftTemplate
    from scheduling.services.eligibility import EligibilityService
    from scheduling.services.engine import shift_window_for

    _generated, show = generated_run(staff_and_config)
    # A bare run: in a generated one he already holds a shift, and would be judged
    # ineligible for clashing with himself rather than for the reason under test.
    run = ScheduleRun.objects.create(
        start_date=show.date, end_date=show.date, status=ScheduleRunStatus.DRAFT
    )
    assistant = EmployeeRole.objects.get(
        employee__display_name="Neil Bobbit", role__name="Bartender"
    )
    assert assistant.on_call_only is True

    floor = EmployeeRole.objects.get(
        employee__display_name="Neil Bobbit", role__name="Server"
    )
    assert floor.active is True
    assert floor.on_call_only is False, "the bar limit must not follow him onto the floor"

    service = EligibilityService()

    def eligible(code, role_name):
        template = ShiftTemplate.objects.get(code=code)
        start, end = shift_window_for(show, template)
        return service.evaluate(
            assistant.employee,
            Role.objects.get(name=role_name),
            show,
            template,
            run,
            start,
            end,
        )

    # He can hold a confirmed floor position...
    assert eligible("lead-server", "Server").eligible is True
    assert eligible("server-3", "Server").eligible is True
    # ...and still cannot hold the bar on his own.
    bar = eligible("bartender", "Bartender")
    assert bar.eligible is False
    assert any("on-call support only" in reason for reason in bar.reasons)
